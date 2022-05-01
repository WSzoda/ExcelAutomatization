import os
import shutil
from openpyxl import *
from openpyxl.styles import *
from openpyxl.drawing.image import Image

def merge_cells(worksheet, low, high):
    worksheet.merge_cells(f'A{low}:A{high}')
    worksheet.merge_cells(f'B{low}:C{high}')
    worksheet.merge_cells(f'D{low}:G{low + 4}')
    worksheet.merge_cells(f'D{low + 5}:G{high - 4}')
    worksheet.merge_cells(f'D{high - 3}:G{high - 2}')
    worksheet.merge_cells(f'D{high - 1}:D{high}')

def add_data(workscheet, low, high, number):
    workscheet[f"A{low}"].value = number
    workscheet[f"D{low+5}"].value = "Zalecenie:"
    workscheet[f"D{high-3}"].value = "Termin wykonania:\nOdpowiedzialny za wykonanie:"
    workscheet[f"D{high-1}"].value = "działanie:"
    workscheet[f"E{high - 1}"].value = "korekcyjne"
    workscheet[f"F{high - 1}"].value = "korygujące"
    workscheet[f"G{high - 1}"].value = "ocena zgodności"
    workscheet[f"G{high}"].value = "........"

def style_cells(workscheet, low, high):
    ft = Font(name="Arial", bold=True, size=8)
    bord = Side(style="thin")
    for i in range(low, high+1):
        workscheet[f"A{i}"].font = ft
        workscheet[f"A{i}"].border = Border(right=bord, left=bord)
        workscheet[f'D{i}'].border = Border(right=bord)
        workscheet[f"B{i}"].font = ft
        workscheet[f"C{i}"].font = ft
        workscheet[f"D{i}"].font = ft
        workscheet[f"E{i}"].font = ft
        workscheet[f"F{i}"].font = ft
        workscheet[f"G{i}"].font = ft
    for i in range(high-1, high+1):
        workscheet[f"D{i}"].alignment = Alignment(horizontal='center', vertical='center')
        workscheet[f"E{i}"].alignment = Alignment(horizontal='center', vertical='center')
        workscheet[f"F{i}"].alignment = Alignment(horizontal='center', vertical='center')
        workscheet[f"G{i}"].alignment = Alignment(horizontal='center', vertical='center')
        workscheet[f"D{i}"].border = Border(outline=True, right=bord, top=bord, left=bord, bottom=bord)
        workscheet[f"E{i}"].border = Border(outline=True, right=bord, top=bord, left=bord, bottom=bord)
        workscheet[f"F{i}"].border = Border(outline=True, right=bord, top=bord, left=bord, bottom=bord)
        workscheet[f"G{i}"].border = Border(outline=True, right=bord, top=bord, left=bord, bottom=bord)
    workscheet[f'A{low}'].border = Border(right=bord, top=bord, left=bord, bottom=bord)
    workscheet[f'D{low+5}'].border = Border(right=bord)
    workscheet[f'A{high}'].border = Border(bottom=bord)
    workscheet[f'D{low + 7}'].border = Border(right=bord)
    workscheet[f'B{low}'].border = Border(top=bord, bottom=bord)
    workscheet[f"D{high-3}"].alignment = Alignment(horizontal='left', vertical='center', wrapText=True)
    workscheet[f"D{high - 5}"].alignment = Alignment(horizontal='left', vertical='center')
    workscheet[f"D{low}"].alignment = Alignment(horizontal='left', vertical='center')
    workscheet[f"A{low}"].alignment = Alignment(horizontal='center', vertical='center')
    workscheet[f"G{high}"].fill = PatternFill("solid", fgColor="00FF0000")
    workscheet[f"G{high}"].alignment = Alignment(horizontal='center', vertical='center')
    workscheet[f"G{high -1 }"].alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

if __name__ == '__main__':
    shutil.copyfile('BHPWzor.xlsx', 'out/wynik.xlsx')
    wb = load_workbook('out/wynik.xlsx')
    ws = wb["Audyt BHP"]
    low = 6
    high = 16
    for i in range(len(os.listdir("zdj"))):
        merge_cells(ws, 6 + 11*i, high + 11*i)
        add_data(ws, 6 + 11*i, high + 11*i, i+1)
        style_cells(ws, 6 + 11*i, high + 11*i)
        img = Image(os.path.join('zdj', os.listdir("zdj")[i]))
        img.height = 245
        img.width = 164
        ws.add_image(img, f'B{6 + 11*i}')
    wb.save("out/wynik.xlsx")

